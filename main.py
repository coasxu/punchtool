import os
import re
import docx # python-docx
import openpyxl

"""
    Author: coasxu
    Version: v0.1
    Update time: 2020/6/7 09:27
"""

# add: judge whether a student has quit punch.
# file change: 名单.xlsx add a columen "未打卡次数"

def cmp(x):
    return x[0]


if __name__ == "__main__":
    # 获取当年目录下唯一的xlsx文件
    xlpath = "【打卡】名单.xlsx"


    print("已加载文件 '%s'" % xlpath)
    # 读取xlsx
    wb=openpyxl.load_workbook(xlpath)
    ws = wb.active

    peoplelist = []
    i = 0
    while True:
        row = 2+i
        nid = ws['A%d'% row].value
        nname = ws['B%d'% row].value
        wname = ws['C%d'% row].value
        loseCount = ws['D%d'% row].value
    #     print(int(nid), nname, wname)
        if wname is None:
            break
        
        loseCount = int(loseCount)
        if loseCount < 3:
            peoplelist.append([int(nid), nname, wname, row])
        i+=1

    peoplelist = sorted(peoplelist, key=cmp)

    path = "【打卡】聊天记录.docx"

    print("已加载文件 '%s'" % path)
    file=docx.Document(path)


    # read chat recording file and extract username and according speaking.
    infos = {}
    usernow = None
    for para in file.paragraphs:
        if usernow is None and para.text != "":
            usernow = para.text.split(":")[0]
            if usernow not in infos:
                infos[usernow] = []
        elif usernow is not None:
            if para.text == "":
                usernow = None
            else:
                infos[usernow].append(para.text)

    co = re.compile(u'[\U00010000-\U0010ffff]')

    # namelist and infolist matching, generate successlist
    successlist = []
    student_fail_indexes = list(range(len(peoplelist)))
    record_faillist = []
    for wname,info in infos.items():
        peop = None
        for index in range(len(peoplelist)):
            people = peoplelist[index]
            clear_wname = co.sub('', wname)
            if co.sub('', people[2]) == clear_wname and clear_wname != '':
                peop = people.copy()
                student_fail_indexes.remove(index)
        if peop is None:
            record_faillist.append([wname, info])
        else:
            peop.extend([info])
            successlist.append(peop)
    # successlist: [student_id, name, wexin_name, row_num, chat_records[]]
    # record_faillist: [wexin_name, chat_records[]]
    
    kickcardlist = []
    # 输出打卡记录
    for line in successlist:
        shougong = 0
        kaigong = 0
        student_id, name, wname, row_num, chat_records = line[0], line[1], line[2], line[3], line[4]
        for ff in chat_records:
            if "开工" in ff:
                kaigong = 1
            if "收工" in ff:
                shougong = 1
        kickcardlist.append([row_num, student_id, name, kaigong, shougong])

    print("******************成功匹配名单******************")
    print("共%d人" % len(kickcardlist))

    print("学号 姓名 开工 收工")
    for line in kickcardlist:
        print(line[1],end=" ")
        print(line[2],end=" ")
        print(line[3],end=" ")
        print(line[4],end=" ")
        print()

    print("******************没有统计到微信名称的名单*****************************")
    print("学号 昵称 微信名称")
    for index in student_fail_indexes:
        print(peoplelist[index][0],end=" ")
        print(peoplelist[index][1],end=" ")
        print(peoplelist[index][2],end=" ")
        print()

    print("******************无法匹配的聊天记录******************")
    print("复制后的名称 聊天记录")
    for line in record_faillist:
        print(line[0],end=" ")
        print(line[1],end=" ")
        print()
    print()
    print("开始修改文件 '%s'" % xlpath)
    # 读取xlsx
    wb=openpyxl.load_workbook(xlpath)
    ws=wb.worksheets[0]
    for card in kickcardlist:
        row = card[0]
        s_id = card[1]
        if card[3] == 1:
            ws["E%d" % row] = 1
        if card[4] == 1:
            ws["F%d" % row] = 1
    wb.save(filename=xlpath)

    print("修改文件 '%s' 完成，已保存！" % xlpath)
    print("各位辛苦啦！")

    line = input("输入任意字符退出")